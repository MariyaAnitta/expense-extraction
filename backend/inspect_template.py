import openpyxl

path = r"c:\Users\AnittaShaji\Downloads\expense extraction\backend\petty_cash_template.xlsx"
wb = openpyxl.load_workbook(path)
ws = wb.active

print("Headers in Row 4:")
for col in range(1, 10):
    val = ws.cell(row=4, column=col).value
    print(f"Col {col}: {val}")

# Check for merged cells or styles in the title area
print("\nFirst row formatting:")
title_cell = ws['A1']
print(f"A1 Value: {title_cell.value}")
print(f"A1 Font: {title_cell.font.name}, {title_cell.font.size}, {title_cell.font.color.rgb if title_cell.font.color else 'None'}")

# Check balance formula
print("\nBalance formulas in Row 5 and 6:")
print(f"F5: {ws.cell(row=5, column=6).value}")
print(f"F6: {ws.cell(row=6, column=6).value}")
