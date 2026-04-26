import openpyxl
import os
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from models import ReceiptData, ExtractionResult
from typing import List

def generate_petty_cash_log(results: List[ExtractionResult], output_path: str, currency: str = "BHD"):
    template_path = os.path.join(os.path.dirname(__file__), "petty_cash_template.xlsx")
    if os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # V2 Step D: Programmatic Find & Replace BHD with target currency
        if currency.strip().upper() != "BHD":
            target_curr = currency.strip().upper()
            for row in ws.iter_rows():
                for cell in row:
                    # 1. Surgical string replacement in cell values (headers/titles)
                    if cell.value and isinstance(cell.value, str) and "BHD" in cell.value:
                        cell.value = cell.value.replace("BHD", target_curr)
                    
                    # 2. Surgical number_format replacement (Corruption fix)
                    # Use currency code directly without extra quotes to avoid visual noise
                    if cell.number_format and "BHD" in cell.number_format:
                        cell.number_format = cell.number_format.replace("BHD", target_curr)
    else:
        # Fallback if template is missing
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Petty Cash Log"

    currency_format = f'{currency} #,##0.000'

    # --- Date Sorting Logic ---
    def parse_date(date_str):
        if not date_str: return datetime.max
        try:
            return datetime.strptime(date_str, "%d/%m/%Y")
        except:
            try:
                return datetime.strptime(date_str, "%Y-%m-%d")
            except:
                return datetime.max

    # Filter and Sort results
    # Priority: 1. Opening Balance (always top), 2. Date
    results_to_process = [r for r in results if r.data]
    results_to_process.sort(key=lambda x: (
        0 if (x.data.description and "opening balance" in str(x.data.description).lower()) else 1,
        parse_date(x.data.date)
    ))

    # --- Determine the year and date range from the receipt dates ---
    all_parsed_dates = []
    for r in results_to_process:
        if r.data and r.data.date:
            parsed = parse_date(r.data.date)
            if parsed != datetime.max:
                all_parsed_dates.append(parsed)

    if all_parsed_dates:
        min_date = min(all_parsed_dates)
        max_date = max(all_parsed_dates)
        receipt_year = min_date.year
    else:
        min_date = None
        max_date = None
        receipt_year = datetime.now().year  # Fallback if no valid dates

    # 1. Update Headers
    try:
        ws['H1'] = receipt_year
        if min_date and max_date:
            ws['A3'] = f"For {min_date.strftime('%d/%m/%Y')} through {max_date.strftime('%d/%m/%Y')}"
        else:
            ws['A3'] = f"For 01/01/{receipt_year} through 31/12/{receipt_year}"
    except: pass

    def safe_float(val):
        if val is None or val == "" or val == "null": return 0.0
        try:
            # Handle cases where currency strings might have commas or extra labels
            s = str(val).replace(currency, '').replace('BHD', '').replace(',', '').strip()
            return float(s)
        except: return 0.0

    # Define formatting utilities at the top to avoid scope/NameErrors
    def get_accounting_fmt(curr_code):
        # Professional Accounting: Currency on left (* fills space), number on right.
        # Semicolon handles negative sign positioning (INR  -1,000.000)
        c = str(curr_code).upper()
        return f'[$ {c}]* #,##0.000;[$ {c}]* -#,##0.000'

    # 2. Total Balance (F3) - Must use functional (normalized) amounts
    current_balance = 0
    for r in results_to_process:
        if not r.data: continue
        d = r.data
        # Use functional_amount (normalized to base currency)
        f_amt = safe_float(d.functional_amount if d.functional_amount is not None else d.base_amount)
        if d.category == "Deposit":
            current_balance += f_amt
        else:
            current_balance -= f_amt
            
    try:
        c_tot = ws['F3']
        c_tot.value = current_balance
        c_tot.number_format = get_accounting_fmt(str(currency).upper())
    except Exception as e:
        print(f"ERROR: Failed to update F3 balance: {e}")

    # 3. Data Rows (Starts at Row 5)
    start_data_row = 5
    row_idx = start_data_row
    running_balance = 0
    
    for result in results_to_process:
        try:
            if not result.data: continue
            d = result.data
            
            # V4.5: Use functional_amount for ledger calculations (Entity Base Currency)
            # Use functional_amount if exists, else fallback to base_amount for legacy data
            f_amt = safe_float(d.functional_amount if d.functional_amount is not None else d.base_amount)
            
            is_dep = (d.category == "Deposit")
            dep_val = f_amt if is_dep else 0
            exp_val = f_amt if not is_dep else 0
            running_balance = running_balance + dep_val - exp_val
            
            orig_amt = safe_float(d.deposit_amount if is_dep else d.amount)
            orig_curr = str(d.currency or currency).upper()

            # DATE (A)
            if d.date:
                try:
                    dt = parse_date(d.date)
                    if dt != datetime.max:
                        c_date = ws.cell(row=row_idx, column=1, value=dt)
                        c_date.number_format = 'yyyy-mm-dd'
                    else:
                        ws.cell(row=row_idx, column=1, value=str(d.date))
                except:
                    ws.cell(row=row_idx, column=1, value=str(d.date))
            
            # DESCRIPTION (B)
            ws.cell(row=row_idx, column=2, value=str(d.description or d.category or "Expense"))
            
            # CATEGORIES (C & D)
            sub_type = d.sub_type or ""
            if is_dep:
                ws.cell(row=row_idx, column=3, value=str(sub_type))
                ws.cell(row=row_idx, column=4, value="")
                # ORIGINAL DEPOSIT (E)
                c_orig = ws.cell(row=row_idx, column=5, value=orig_amt)
                c_orig.number_format = get_accounting_fmt(orig_curr)
                ws.cell(row=row_idx, column=6, value="")
            else:
                ws.cell(row=row_idx, column=3, value="")
                ws.cell(row=row_idx, column=4, value=str(sub_type))
                # ORIGINAL EXPENSE (F)
                ws.cell(row=row_idx, column=5, value="")
                c_orig = ws.cell(row=row_idx, column=6, value=orig_amt)
                c_orig.number_format = get_accounting_fmt(orig_curr)

            # AUDITED AMOUNT & RATE (G & H)
            target_curr = str(d.target_currency or currency).upper()
            if orig_curr != target_curr or d.is_manual_rate:
                t_amt = safe_float(d.base_amount)
                c_audit = ws.cell(row=row_idx, column=7, value=t_amt)
                c_audit.number_format = get_accounting_fmt(target_curr)
                
                c_rate = ws.cell(row=row_idx, column=8, value=safe_float(d.exchange_rate))
                # High precision + Currency label
                c_rate.number_format = f'[$ {target_curr}] * 0.000000'
            else:
                ws.cell(row=row_idx, column=7, value="")
                ws.cell(row=row_idx, column=8, value="")

            # FINAL AMOUNT & RATE (I & J)
            func_curr = str(d.functional_currency or currency).upper()
            
            c_func = ws.cell(row=row_idx, column=9, value=f_amt)
            c_func.number_format = get_accounting_fmt(func_curr)
            
            c_frate = ws.cell(row=row_idx, column=10, value=safe_float(d.functional_rate or d.exchange_rate))
            c_frate.number_format = f'[$ {func_curr}] * 0.000000'
            
            # RECEIVED BY (K)
            ws.cell(row=row_idx, column=11, value=str(d.received_by or ""))
            
            # BALANCE (L)
            c_bal = ws.cell(row=row_idx, column=12, value=running_balance) 
            c_bal.number_format = get_accounting_fmt(func_curr)
            
            # REMARKS (M)
            ws.cell(row=row_idx, column=13, value=str(d.remarks or "ok")) 

            row_idx += 1
        except Exception as e:
            print(f"ERROR: Failed to process row {row_idx}: {e}")
            continue

    # 4. COLUMN WIDTHS (V6 - Fix "####")
    for col_let in ['E', 'F', 'G', 'I', 'L']:
        ws.column_dimensions[col_let].width = 22

    # 5. TEMPLATE TOTALS (V5 Fixed at Row 123)
    try:
        t_row = 123
        func_curr = str(currency).upper() 
        
        # Calculate sums in functional (base) currency to avoid mixed-currency math errors
        total_dep = 0
        total_exp = 0
        for r in results_to_process:
            if not r.data: continue
            val = safe_float(r.data.functional_amount if r.data.functional_amount is not None else r.data.base_amount)
            if r.data.category == "Deposit":
                total_dep += val
            else:
                total_exp += val

        # In Row 123, we prioritize the BASE CURRENCY TOTALS as requested by the user
        ws.cell(row=t_row, column=5, value=total_dep)
        ws.cell(row=t_row, column=6, value=total_exp)
        
        # User requested NO balance/total in Column I for the total row
        ws.cell(row=t_row, column=9, value="")
        
        # Apply bold and accounting format to totals
        for col in [5, 6]:
            c = ws.cell(row=t_row, column=col)
            c.font = Font(bold=True)
            c.number_format = get_accounting_fmt(func_curr)
    except Exception as e:
        print(f"ERROR: Failed to update Template Total row 123: {e}")

    wb.save(output_path)
